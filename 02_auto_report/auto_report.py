import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report(input_path, output_path):
    df = pd.read_excel(input_path)

    # 加總列
    totals = {}
    for col in df.select_dtypes(include="number").columns:
        totals[col] = df[col].sum()
    total_row = {col: totals.get(col, "") for col in df.columns}
    total_row[df.columns[0]] = "Total"
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    df.to_excel(output_path, index=False, sheet_name="Report")

    # 套用樣式
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="2F5496")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=last_row - 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[1].height = 30

    wb.save(output_path)
    print(f"Done! Report saved → {output_path}")

if __name__ == "__main__":
    input_path = "sample_input/data.xlsx"
    output_path = "report_output.xlsx"
    generate_report(input_path, output_path)
